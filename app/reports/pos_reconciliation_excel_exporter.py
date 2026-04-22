from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.core.config import settings
from app.services.pos_report_service import get_pos_reconciliation_report


COLOR_NAVY = "1F4E78"
COLOR_BLUE = "D9EAF7"
COLOR_LIGHT_BLUE = "EAF4FB"
COLOR_LIGHT_GRAY = "F3F6F8"
COLOR_WHITE = "FFFFFF"
COLOR_BLACK = "000000"
COLOR_GREEN = "D9EAD3"
COLOR_DARK_GREEN = "38761D"
COLOR_RED = "F4CCCC"
COLOR_DARK_RED = "990000"
COLOR_YELLOW = "FFF2CC"
COLOR_ORANGE = "FCE5CD"
COLOR_GRAY = "D9D9D9"

MONEY_FORMAT = '#,##0.00'
RATE_FORMAT = '0.000000'
DATE_FORMAT = 'DD.MM.YYYY'

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)

TITLE_FONT = Font(color=COLOR_WHITE, bold=True, size=18)
SUBTITLE_FONT = Font(color=COLOR_WHITE, size=11)
HEADER_FONT = Font(color=COLOR_WHITE, bold=True, size=11)
NORMAL_FONT = Font(color=COLOR_BLACK, size=10)
BOLD_FONT = Font(color=COLOR_BLACK, bold=True, size=10)
CARD_FONT = Font(color=COLOR_WHITE, bold=True, size=14)


def _as_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0.00")

    if isinstance(value, Decimal):
        return value

    return Decimal(str(value))


def _to_excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (date, datetime)):
        return value

    return value


def _format_money_text(value: object) -> str:
    decimal_value = _as_decimal(value)

    return f"{decimal_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]

    cleaned_title = title

    for char in invalid_chars:
        cleaned_title = cleaned_title.replace(char, "-")

    return cleaned_title[:31]


def _style_title(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:L2")
    ws["A2"] = subtitle
    ws["A2"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24


def _style_header_row(ws, row_number: int) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue

        cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_body(ws, start_row: int) -> None:
    if ws.max_row < start_row:
        return

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY)


def _apply_status_style(ws, status_column: int, start_row: int) -> None:
    if ws.max_row < start_row:
        return

    for row_number in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_number, column=status_column)

        if cell.value is None:
            continue

        status = str(cell.value).strip().upper()

        if status == "MISMATCH":
            cell.fill = PatternFill("solid", fgColor=COLOR_RED)
            cell.font = Font(color=COLOR_DARK_RED, bold=True)

        elif status == "REALIZED":
            cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
            cell.font = Font(color=COLOR_DARK_GREEN, bold=True)

        elif status == "PLANNED":
            cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)
            cell.font = Font(color="7F6000", bold=True)

        elif status == "CANCELLED":
            cell.fill = PatternFill("solid", fgColor=COLOR_GRAY)
            cell.font = Font(color=COLOR_BLACK, bold=True)


def _apply_formats(
    ws,
    *,
    money_columns: list[int] | None = None,
    rate_columns: list[int] | None = None,
    date_columns: list[int] | None = None,
    start_row: int = 2,
) -> None:
    money_columns = money_columns or []
    rate_columns = rate_columns or []
    date_columns = date_columns or []

    for row_number in range(start_row, ws.max_row + 1):
        for column_index in money_columns:
            ws.cell(row=row_number, column=column_index).number_format = MONEY_FORMAT

        for column_index in rate_columns:
            ws.cell(row=row_number, column=column_index).number_format = RATE_FORMAT

        for column_index in date_columns:
            ws.cell(row=row_number, column=column_index).number_format = DATE_FORMAT


def _autofit_columns(ws, min_width: int = 10, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            value_length = len(str(cell.value))

            if value_length > max_length:
                max_length = value_length

        ws.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)


def _write_table(
    ws,
    *,
    headers: list[str],
    rows: list[list[Any]],
    money_columns: list[int] | None = None,
    rate_columns: list[int] | None = None,
    date_columns: list[int] | None = None,
    status_column: int | None = None,
) -> None:
    ws.append(headers)

    for row in rows:
        ws.append([_to_excel_value(value) for value in row])

    ws.freeze_panes = "A2"

    _style_header_row(ws, 1)
    _style_body(ws, 2)
    _apply_formats(
        ws,
        money_columns=money_columns,
        rate_columns=rate_columns,
        date_columns=date_columns,
        start_row=2,
    )

    if status_column is not None:
        _apply_status_style(ws, status_column, 2)

    _autofit_columns(ws)


def _write_dashboard_card(ws, cell_range: str, title: str, value: str, color: str) -> None:
    ws.merge_cells(cell_range)

    start_cell = cell_range.split(":")[0]
    cell = ws[start_cell]
    cell.value = f"{title}\n{value}"
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = CARD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _calculate_dashboard_totals(report: dict[str, Any]) -> dict[str, Decimal | int]:
    total_gross_amount = Decimal("0.00")
    total_commission_amount = Decimal("0.00")

    pending_expected_net_amount = Decimal("0.00")

    realized_expected_net_amount = Decimal("0.00")
    actual_net_amount = Decimal("0.00")
    reconciliation_difference_amount = Decimal("0.00")

    mismatch_count = 0
    planned_count = 0
    realized_count = 0

    for row in report["detail_rows"]:
        status = str(row["status"]).strip().upper()

        gross_amount = _as_decimal(row["gross_amount"])
        commission_amount = _as_decimal(row["commission_amount"])
        expected_net_amount = _as_decimal(row["expected_net_amount"])
        row_actual_net_amount = _as_decimal(row["actual_net_amount"])
        difference_amount = _as_decimal(row["difference_amount"])

        total_gross_amount += gross_amount
        total_commission_amount += commission_amount

        if status == "PLANNED":
            planned_count += 1
            pending_expected_net_amount += expected_net_amount

        elif status in {"REALIZED", "MISMATCH"}:
            realized_count += 1
            realized_expected_net_amount += expected_net_amount
            actual_net_amount += row_actual_net_amount
            reconciliation_difference_amount += difference_amount

            if status == "MISMATCH":
                mismatch_count += 1

    return {
        "total_gross_amount": total_gross_amount,
        "total_commission_amount": total_commission_amount,
        "pending_expected_net_amount": pending_expected_net_amount,
        "realized_expected_net_amount": realized_expected_net_amount,
        "actual_net_amount": actual_net_amount,
        "reconciliation_difference_amount": reconciliation_difference_amount,
        "mismatch_count": mismatch_count,
        "planned_count": planned_count,
        "realized_count": realized_count,
    }


def _create_dashboard_sheet(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Dashboard"))

    start_date = report["start_date"]
    end_date = report["end_date"]
    mismatch_rows = report["mismatch_rows"]

    dashboard_totals = _calculate_dashboard_totals(report)

    _style_title(
        ws,
        "FTM - POS Mutabakat Raporu",
        (
            f"Rapor dönemi: {start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')} | "
            "Bekleyen ve gerçekleşen POS tutarları ayrı hesaplanır"
        ),
    )

    _write_dashboard_card(
        ws,
        "A4:B6",
        "BRÜT POS TOPLAMI",
        f"{_format_money_text(dashboard_totals['total_gross_amount'])} TRY",
        COLOR_NAVY,
    )

    _write_dashboard_card(
        ws,
        "C4:D6",
        "BEKLEYEN NET POS",
        f"{_format_money_text(dashboard_totals['pending_expected_net_amount'])} TRY",
        "C65911",
    )

    _write_dashboard_card(
        ws,
        "E4:F6",
        "GERÇEKLEŞEN BEKLENEN NET",
        f"{_format_money_text(dashboard_totals['realized_expected_net_amount'])} TRY",
        COLOR_DARK_GREEN,
    )

    _write_dashboard_card(
        ws,
        "G4:H6",
        "GERÇEK YATAN",
        f"{_format_money_text(dashboard_totals['actual_net_amount'])} TRY",
        COLOR_NAVY,
    )

    _write_dashboard_card(
        ws,
        "I4:J6",
        "MUTABAKAT FARKI",
        f"{_format_money_text(dashboard_totals['reconciliation_difference_amount'])} TRY",
        COLOR_DARK_RED
        if dashboard_totals["reconciliation_difference_amount"] != Decimal("0.00")
        else COLOR_DARK_GREEN,
    )

    _write_dashboard_card(
        ws,
        "K4:L6",
        "MISMATCH ADEDİ",
        str(dashboard_totals["mismatch_count"]),
        COLOR_DARK_RED if dashboard_totals["mismatch_count"] else COLOR_DARK_GREEN,
    )

    ws["A8"] = "Dashboard Açıklaması"
    ws["A8"].font = Font(bold=True, size=13, color=COLOR_NAVY)

    explanation_rows = [
        ["Brüt POS Toplamı", "Seçilen tarih aralığındaki tüm POS brüt satışlarının toplamıdır."],
        ["Bekleyen Net POS", "Henüz bankaya yatışı gerçekleşmemiş PLANNED kayıtların beklenen net toplamıdır."],
        ["Gerçekleşen Beklenen Net", "REALIZED ve MISMATCH kayıtların sistemde beklenen net toplamıdır."],
        ["Gerçek Yatan", "REALIZED ve MISMATCH kayıtlar için bankaya fiilen yatan toplamdır."],
        ["Mutabakat Farkı", "Sadece gerçekleşmiş kayıtlar üzerinden hesaplanır. Gerçek yatan - gerçekleşen beklenen net."],
        ["MISMATCH Adedi", "Beklenen net ile gerçek yatan tutarı farklı olan POS kayıt sayısıdır."],
    ]

    row_number = 9

    for title, description in explanation_rows:
        ws.cell(row=row_number, column=1).value = title
        ws.cell(row=row_number, column=1).font = BOLD_FONT
        ws.cell(row=row_number, column=1).fill = PatternFill("solid", fgColor=COLOR_LIGHT_BLUE)
        ws.cell(row=row_number, column=1).border = THIN_BORDER

        ws.merge_cells(start_row=row_number, start_column=2, end_row=row_number, end_column=12)
        ws.cell(row=row_number, column=2).value = description
        ws.cell(row=row_number, column=2).font = NORMAL_FONT
        ws.cell(row=row_number, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row_number, column=2).border = THIN_BORDER

        row_number += 1

    mismatch_title_row = row_number + 2

    ws.cell(row=mismatch_title_row, column=1).value = "MISMATCH Kayıtları"
    ws.cell(row=mismatch_title_row, column=1).font = Font(bold=True, size=13, color=COLOR_NAVY)

    headers = [
        "ID",
        "POS",
        "Banka / Hesap",
        "İşlem Tarihi",
        "Beklenen Yatış",
        "Gerçekleşen",
        "Beklenen Net",
        "Gerçek Net",
        "Fark",
        "Fark Nedeni",
        "Durum",
        "Banka Hareket ID",
    ]

    header_row = mismatch_title_row + 1

    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=column_index)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    data_row = header_row + 1

    for row in mismatch_rows:
        values = [
            row["id"],
            row["pos_label"],
            row["bank_label"],
            row["transaction_date"],
            row["expected_settlement_date"],
            row["realized_settlement_date"],
            row["expected_net_amount"],
            row["actual_net_amount"],
            row["difference_amount"],
            row["difference_reason"],
            row["status"],
            row["bank_transaction_id"],
        ]

        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=column_index)
            cell.value = _to_excel_value(value)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            if column_index in [7, 8, 9]:
                cell.number_format = MONEY_FORMAT

            if column_index in [4, 5, 6]:
                cell.number_format = DATE_FORMAT

        data_row += 1

    _apply_status_style(ws, 11, header_row + 1)

    ws.freeze_panes = f"A{header_row}"

    for column_letter in ["A", "D", "E", "F", "G", "H", "I", "K", "L"]:
        ws.column_dimensions[column_letter].width = 16

    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["J"].width = 40


def _create_detail_sheet(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("POS Detayları"))

    rows = []

    for row in report["detail_rows"]:
        rows.append(
            [
                row["id"],
                row["pos_label"],
                row["bank_label"],
                row["transaction_date"],
                row["expected_settlement_date"],
                row["realized_settlement_date"],
                row["gross_amount"],
                row["commission_rate"],
                row["commission_amount"],
                row["expected_net_amount"],
                row["actual_net_amount"],
                row["difference_amount"],
                row["difference_reason"],
                row["currency_code"],
                row["status"],
                row["bank_transaction_id"],
                row["reference_no"],
                row["description"],
            ]
        )

    _write_table(
        ws,
        headers=[
            "ID",
            "POS",
            "Banka / Hesap",
            "İşlem Tarihi",
            "Beklenen Yatış",
            "Gerçekleşen",
            "Brüt Tutar",
            "Komisyon Oranı",
            "Komisyon Tutarı",
            "Beklenen Net",
            "Gerçek Net",
            "Fark",
            "Fark Nedeni",
            "Para Birimi",
            "Durum",
            "Banka Hareket ID",
            "Referans No",
            "Açıklama",
        ],
        rows=rows,
        money_columns=[7, 9, 10, 11, 12],
        rate_columns=[8],
        date_columns=[4, 5, 6],
        status_column=15,
    )


def _create_mismatch_sheet(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("MISMATCH Kayıtları"))

    rows = []

    for row in report["mismatch_rows"]:
        rows.append(
            [
                row["id"],
                row["pos_label"],
                row["bank_label"],
                row["transaction_date"],
                row["expected_settlement_date"],
                row["realized_settlement_date"],
                row["expected_net_amount"],
                row["actual_net_amount"],
                row["difference_amount"],
                row["difference_reason"],
                row["currency_code"],
                row["bank_transaction_id"],
                row["reference_no"],
                row["description"],
            ]
        )

    _write_table(
        ws,
        headers=[
            "ID",
            "POS",
            "Banka / Hesap",
            "İşlem Tarihi",
            "Beklenen Yatış",
            "Gerçekleşen",
            "Beklenen Net",
            "Gerçek Net",
            "Fark",
            "Fark Nedeni",
            "Para Birimi",
            "Banka Hareket ID",
            "Referans No",
            "Açıklama",
        ],
        rows=rows,
        money_columns=[7, 8, 9],
        date_columns=[4, 5, 6],
    )


def _create_status_summary_sheet(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Durum Özeti"))

    rows = []

    for status, totals in report["totals_by_status"].items():
        rows.append(
            [
                status,
                totals["gross_total"],
                totals["commission_total"],
                totals["expected_net_total"],
                totals["actual_net_total"],
                totals["difference_total"],
            ]
        )

    _write_table(
        ws,
        headers=[
            "Durum",
            "Brüt Toplam",
            "Komisyon Toplamı",
            "Beklenen Net Toplam",
            "Gerçek Yatan Toplam",
            "Fark Toplamı",
        ],
        rows=rows,
        money_columns=[2, 3, 4, 5, 6],
        status_column=1,
    )


def _create_bank_summary_sheet(wb: Workbook, report: dict[str, Any]) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Banka Özeti"))

    rows = []

    for bank_label, totals in report["totals_by_bank"].items():
        rows.append(
            [
                bank_label,
                totals["gross_total"],
                totals["commission_total"],
                totals["expected_net_total"],
                totals["actual_net_total"],
                totals["difference_total"],
            ]
        )

    _write_table(
        ws,
        headers=[
            "Banka / Hesap",
            "Brüt Toplam",
            "Komisyon Toplamı",
            "Beklenen Net Toplam",
            "Gerçek Yatan Toplam",
            "Fark Toplamı",
        ],
        rows=rows,
        money_columns=[2, 3, 4, 5, 6],
    )


def export_pos_reconciliation_to_excel(
    session: Session,
    *,
    start_date: date,
    end_date: date,
) -> Path:
    settings.export_folder.mkdir(parents=True, exist_ok=True)

    report = get_pos_reconciliation_report(
        session,
        start_date=start_date,
        end_date=end_date,
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = settings.export_folder / f"pos_mutabakat_raporu_{timestamp}.xlsx"

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    _create_dashboard_sheet(wb, report)
    _create_detail_sheet(wb, report)
    _create_mismatch_sheet(wb, report)
    _create_status_summary_sheet(wb, report)
    _create_bank_summary_sheet(wb, report)

    wb.active = 0
    wb.save(output_path)

    return output_path