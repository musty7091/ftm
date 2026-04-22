from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, aliased

from app.core.config import settings
from app.models.bank import Bank, BankAccount
from app.models.business_partner import BusinessPartner
from app.models.check import IssuedCheck, ReceivedCheck
from app.services.bank_transaction_service import get_bank_account_balance_summary
from app.services.risk_service import get_all_bank_risk_summaries
from app.services.transfer_recommendation_service import get_all_transfer_recommendations


COLOR_NAVY = "1F4E78"
COLOR_BLUE = "D9EAF7"
COLOR_LIGHT_GRAY = "F3F6F8"
COLOR_WHITE = "FFFFFF"
COLOR_GREEN = "D9EAD3"
COLOR_DARK_GREEN = "38761D"
COLOR_RED = "F4CCCC"
COLOR_DARK_RED = "990000"
COLOR_YELLOW = "FFF2CC"
COLOR_ORANGE = "FCE5CD"
COLOR_GRAY = "D9D9D9"
COLOR_BLACK = "000000"

MONEY_FORMAT = '#,##0.00'
DATE_FORMAT = 'DD.MM.YYYY'

THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2EC"),
    right=Side(style="thin", color="D9E2EC"),
    top=Side(style="thin", color="D9E2EC"),
    bottom=Side(style="thin", color="D9E2EC"),
)

TITLE_FONT = Font(color=COLOR_WHITE, bold=True, size=18)
SUBTITLE_FONT = Font(color=COLOR_WHITE, size=11)
HEADER_FONT = Font(color=COLOR_WHITE, bold=True, size=11)
NORMAL_FONT = Font(color=COLOR_BLACK, size=10)
BOLD_FONT = Font(color=COLOR_BLACK, bold=True, size=10)
CARD_FONT = Font(color=COLOR_WHITE, bold=True, size=15)


def _safe_sheet_title(title: str) -> str:
    invalid_chars = ["\\", "/", "*", "[", "]", ":", "?"]

    cleaned_title = title

    for char in invalid_chars:
        cleaned_title = cleaned_title.replace(char, "-")

    return cleaned_title[:31]


def _as_decimal(value: object) -> Decimal:
    if value is None:
        return Decimal("0.00")

    if isinstance(value, Decimal):
        return value

    return Decimal(str(value))


def _to_excel_value(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)

    if isinstance(value, (date, datetime)):
        return value

    return value


def _format_money_text(value: object) -> str:
    decimal_value = _as_decimal(value)

    return f"{decimal_value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _style_header_row(ws, row_number: int = 1) -> None:
    for cell in ws[row_number]:
        if cell.value is None:
            continue

        cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _apply_body_style(ws, start_row: int = 2) -> None:
    if ws.max_row < start_row:
        return

    for row in ws.iter_rows(min_row=start_row):
        for cell in row:
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOR_LIGHT_GRAY)


def _apply_column_formats(ws, money_columns: list[int], date_columns: list[int], start_row: int = 2) -> None:
    if ws.max_row < start_row:
        return

    for column_index in money_columns:
        for row_number in range(start_row, ws.max_row + 1):
            ws.cell(row=row_number, column=column_index).number_format = MONEY_FORMAT

    for column_index in date_columns:
        for row_number in range(start_row, ws.max_row + 1):
            ws.cell(row=row_number, column=column_index).number_format = DATE_FORMAT


def _apply_status_colors(ws, status_column: int, start_row: int = 2) -> None:
    if ws.max_row < start_row:
        return

    for row_number in range(start_row, ws.max_row + 1):
        cell = ws.cell(row=row_number, column=status_column)

        if cell.value is None:
            continue

        status = str(cell.value).strip().upper()

        if status in {"RISK", "KAPANMAYAN_RİSK"}:
            cell.fill = PatternFill("solid", fgColor=COLOR_RED)
            cell.font = Font(color=COLOR_DARK_RED, bold=True)

        elif status in {"TAKIP", "GIVEN", "PREPARED", "PORTFOLIO", "IN_COLLECTION", "GIVEN_TO_BANK"}:
            cell.fill = PatternFill("solid", fgColor=COLOR_YELLOW)
            cell.font = Font(color="7F6000", bold=True)

        elif status in {"OK", "PAID", "COLLECTED", "AKTIF", "REALIZED"}:
            cell.fill = PatternFill("solid", fgColor=COLOR_GREEN)
            cell.font = Font(color=COLOR_DARK_GREEN, bold=True)

        elif status in {"CANCELLED", "PASIF"}:
            cell.fill = PatternFill("solid", fgColor=COLOR_GRAY)
            cell.font = Font(color=COLOR_BLACK, bold=True)

        elif status in {"TRANSFER_ÖNERİSİ", "KALAN_FAZLA"}:
            cell.fill = PatternFill("solid", fgColor=COLOR_ORANGE)
            cell.font = Font(color="7F6000", bold=True)


def _autofit_columns(ws, min_width: int = 10, max_width: int = 45) -> None:
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = 0

        for cell in column_cells:
            if cell.value is None:
                continue

            value_length = len(str(cell.value))

            if value_length > max_length:
                max_length = value_length

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def _write_table(
    ws,
    headers: list[str],
    rows: list[list[Any]],
    *,
    money_columns: list[int] | None = None,
    date_columns: list[int] | None = None,
    status_column: int | None = None,
) -> None:
    money_columns = money_columns or []
    date_columns = date_columns or []

    ws.append(headers)

    for row in rows:
        ws.append([_to_excel_value(value) for value in row])

    ws.freeze_panes = "A2"

    _style_header_row(ws, 1)
    _apply_body_style(ws, 2)
    _apply_column_formats(ws, money_columns, date_columns)

    if status_column is not None:
        _apply_status_colors(ws, status_column)

    _autofit_columns(ws)


def _create_dashboard_card(ws, cell_range: str, title: str, value: str, color: str) -> None:
    ws.merge_cells(cell_range)

    start_cell = cell_range.split(":")[0]
    cell = ws[start_cell]
    cell.value = f"{title}\n{value}"
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = CARD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def _sum_by_try(current_value: Decimal, amount: object) -> Decimal:
    return current_value + _as_decimal(amount)


def _create_dashboard_sheet(wb: Workbook, session: Session, as_of_date: date) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Dashboard"))

    ws.merge_cells("A1:H1")
    ws["A1"] = "FTM - Finansal Takip Merkezi"
    ws["A1"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Finansal özet raporu | Rapor tarihi: {as_of_date.strftime('%d.%m.%Y')}"
    ws["A2"].fill = PatternFill("solid", fgColor=COLOR_NAVY)
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    total_bank_balance = Decimal("0.00")
    total_pending_issued = Decimal("0.00")
    total_pending_received = Decimal("0.00")
    total_risk_open = Decimal("0.00")

    bank_accounts = session.execute(
        select(BankAccount).where(BankAccount.is_active.is_(True)).order_by(BankAccount.id)
    ).scalars().all()

    for bank_account in bank_accounts:
        summary = get_bank_account_balance_summary(session, bank_account_id=bank_account.id)

        if summary["currency_code"] == "TRY":
            total_bank_balance = _sum_by_try(total_bank_balance, summary["current_balance"])

    all_risks = get_all_bank_risk_summaries(session, as_of_date=as_of_date)
    risk_30_rows = all_risks.get(30, [])

    for row in risk_30_rows:
        if row["currency_code"] == "TRY":
            total_pending_issued = _sum_by_try(total_pending_issued, row["pending_issued_checks_total"])
            total_pending_received = _sum_by_try(total_pending_received, row["expected_received_checks_total"])

            projected_balance = _as_decimal(row["projected_balance"])

            if projected_balance < Decimal("0.00"):
                total_risk_open += abs(projected_balance)

    _create_dashboard_card(
        ws,
        "A4:B6",
        "TOPLAM BANKA BAKİYESİ",
        f"{_format_money_text(total_bank_balance)} TRY",
        COLOR_NAVY,
    )

    _create_dashboard_card(
        ws,
        "C4:D6",
        "BEKLEYEN ÇEK YÜKÜ",
        f"{_format_money_text(total_pending_issued)} TRY",
        "C65911",
    )

    _create_dashboard_card(
        ws,
        "E4:F6",
        "BEKLENEN MÜŞTERİ ÇEKİ",
        f"{_format_money_text(total_pending_received)} TRY",
        COLOR_DARK_GREEN,
    )

    _create_dashboard_card(
        ws,
        "G4:H6",
        "30 GÜNLÜK RİSK AÇIĞI",
        f"{_format_money_text(total_risk_open)} TRY",
        COLOR_DARK_RED if total_risk_open > Decimal("0.00") else COLOR_DARK_GREEN,
    )

    ws["A8"] = "30 Günlük Banka Risk Özeti"
    ws["A8"].font = Font(bold=True, size=13, color=COLOR_NAVY)

    headers = [
        "Banka",
        "Hesap",
        "Para Birimi",
        "Güncel Bakiye",
        "Yazdığımız Çekler",
        "Beklenen Müşteri Çeki",
        "Tahmini Bakiye",
        "Durum",
    ]

    for column_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=column_index)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    current_row = 10

    for risk in risk_30_rows:
        values = [
            risk["bank_name"],
            risk["account_name"],
            risk["currency_code"],
            risk["current_balance"],
            risk["pending_issued_checks_total"],
            risk["expected_received_checks_total"],
            risk["projected_balance"],
            risk["risk_status"],
        ]

        for column_index, value in enumerate(values, start=1):
            cell = ws.cell(row=current_row, column=column_index)
            cell.value = _to_excel_value(value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = NORMAL_FONT

            if column_index in [4, 5, 6, 7]:
                cell.number_format = MONEY_FORMAT

        current_row += 1

    _apply_status_colors(ws, 8, 10)

    ws["A14"] = "30 Günlük Transfer / Aksiyon Önerileri"
    ws["A14"].font = Font(bold=True, size=13, color=COLOR_NAVY)

    recommendation_headers = [
        "Kayıt Türü",
        "Çıkış Hesabı",
        "Giriş / Risk Hesabı",
        "Para Birimi",
        "Tutar",
        "Açıklama",
    ]

    header_row = 15

    for column_index, header in enumerate(recommendation_headers, start=1):
        cell = ws.cell(row=header_row, column=column_index)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    recommendations = get_all_transfer_recommendations(session, as_of_date=as_of_date).get(30)
    data_row = 16

    if recommendations:
        for recommendation in recommendations["recommendations"]:
            values = [
                "TRANSFER_ÖNERİSİ",
                recommendation["from_account_label"],
                recommendation["to_account_label"],
                recommendation["currency_code"],
                recommendation["amount"],
                recommendation["reason"],
            ]

            for column_index, value in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=column_index)
                cell.value = _to_excel_value(value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if column_index == 5:
                    cell.number_format = MONEY_FORMAT

            data_row += 1

        for risk in recommendations["unresolved_risks"]:
            values = [
                "KAPANMAYAN_RİSK",
                "",
                risk["account_label"],
                risk["currency_code"],
                risk["remaining_need"],
                "Dış kaynak, nakit yatırma veya ek tahsilat gerekiyor.",
            ]

            for column_index, value in enumerate(values, start=1):
                cell = ws.cell(row=data_row, column=column_index)
                cell.value = _to_excel_value(value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

                if column_index == 5:
                    cell.number_format = MONEY_FORMAT

            data_row += 1

    _apply_status_colors(ws, 1, 16)

    for column_letter in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[column_letter].width = 24

    ws.column_dimensions["F"].width = 45
    ws.freeze_panes = "A9"


def _create_bank_balances_sheet(wb: Workbook, session: Session) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Banka Bakiyeleri"))

    rows = []

    statement = (
        select(BankAccount, Bank)
        .join(Bank, BankAccount.bank_id == Bank.id)
        .order_by(Bank.name, BankAccount.account_name)
    )

    for bank_account, bank in session.execute(statement).all():
        summary = get_bank_account_balance_summary(session, bank_account_id=bank_account.id)

        rows.append(
            [
                bank.name,
                bank_account.account_name,
                summary["currency_code"],
                summary["opening_balance"],
                summary["incoming_total"],
                summary["outgoing_total"],
                summary["current_balance"],
                "AKTIF" if bank_account.is_active else "PASIF",
            ]
        )

    _write_table(
        ws,
        [
            "Banka",
            "Hesap",
            "Para Birimi",
            "Açılış Bakiyesi",
            "Toplam Giriş",
            "Toplam Çıkış",
            "Güncel Bakiye",
            "Durum",
        ],
        rows,
        money_columns=[4, 5, 6, 7],
        status_column=8,
    )


def _create_risk_sheets(wb: Workbook, session: Session, as_of_date: date) -> None:
    all_risks = get_all_bank_risk_summaries(session, as_of_date=as_of_date)

    for horizon_days, summaries in all_risks.items():
        ws = wb.create_sheet(_safe_sheet_title(f"{horizon_days} Günlük Risk"))

        rows = []

        for summary in summaries:
            rows.append(
                [
                    summary["as_of_date"],
                    summary["cutoff_date"],
                    summary["bank_name"],
                    summary["account_name"],
                    summary["currency_code"],
                    summary["current_balance"],
                    summary["pending_issued_checks_total"],
                    summary["expected_received_checks_total"],
                    summary["projected_balance"],
                    summary["risk_status"],
                ]
            )

        _write_table(
            ws,
            [
                "Rapor Tarihi",
                "Kesim Tarihi",
                "Banka",
                "Hesap",
                "Para Birimi",
                "Güncel Bakiye",
                "Yazdığımız Çekler",
                "Beklenen Müşteri Çeki",
                "Tahmini Bakiye",
                "Risk Durumu",
            ],
            rows,
            money_columns=[6, 7, 8, 9],
            date_columns=[1, 2],
            status_column=10,
        )


def _create_transfer_recommendations_sheet(wb: Workbook, session: Session, as_of_date: date) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Transfer Önerileri"))

    all_reports = get_all_transfer_recommendations(session, as_of_date=as_of_date)

    rows = []

    for horizon_days, report in all_reports.items():
        for recommendation in report["recommendations"]:
            rows.append(
                [
                    horizon_days,
                    "TRANSFER_ÖNERİSİ",
                    recommendation["from_account_label"],
                    recommendation["to_account_label"],
                    recommendation["currency_code"],
                    recommendation["amount"],
                    recommendation["reason"],
                ]
            )

        for risk in report["unresolved_risks"]:
            rows.append(
                [
                    horizon_days,
                    "KAPANMAYAN_RİSK",
                    "",
                    risk["account_label"],
                    risk["currency_code"],
                    risk["remaining_need"],
                    "Dış kaynak, nakit yatırma veya ek tahsilat gerekiyor.",
                ]
            )

        for surplus in report["unused_surpluses"]:
            rows.append(
                [
                    horizon_days,
                    "KALAN_FAZLA",
                    surplus["account_label"],
                    "",
                    surplus["currency_code"],
                    surplus["available_amount"],
                    "Transfer sonrası kullanılabilir fazla.",
                ]
            )

    _write_table(
        ws,
        [
            "Gün",
            "Kayıt Türü",
            "Çıkış Hesabı",
            "Giriş / Risk Hesabı",
            "Para Birimi",
            "Tutar",
            "Açıklama",
        ],
        rows,
        money_columns=[6],
        status_column=2,
    )


def _create_issued_checks_sheet(wb: Workbook, session: Session) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Yazdığımız Çekler"))

    rows = []

    statement = (
        select(IssuedCheck, BusinessPartner, BankAccount, Bank)
        .join(BusinessPartner, IssuedCheck.supplier_id == BusinessPartner.id)
        .join(BankAccount, IssuedCheck.bank_account_id == BankAccount.id)
        .join(Bank, BankAccount.bank_id == Bank.id)
        .order_by(IssuedCheck.due_date, IssuedCheck.id)
    )

    for check, supplier, bank_account, bank in session.execute(statement).all():
        rows.append(
            [
                check.id,
                check.check_number,
                supplier.name,
                bank.name,
                bank_account.account_name,
                check.issue_date,
                check.due_date,
                check.amount,
                check.currency_code.value,
                check.status.value,
                check.paid_transaction_id,
                check.reference_no,
                check.description,
            ]
        )

    _write_table(
        ws,
        [
            "ID",
            "Çek No",
            "Tedarikçi",
            "Banka",
            "Hesap",
            "Düzenleme Tarihi",
            "Vade Tarihi",
            "Tutar",
            "Para Birimi",
            "Durum",
            "Ödeme Hareket ID",
            "Referans No",
            "Açıklama",
        ],
        rows,
        money_columns=[8],
        date_columns=[6, 7],
        status_column=10,
    )


def _create_received_checks_sheet(wb: Workbook, session: Session) -> None:
    ws = wb.create_sheet(_safe_sheet_title("Aldığımız Çekler"))

    rows = []

    collection_account_alias = aliased(BankAccount)
    collection_bank_alias = aliased(Bank)

    statement = (
        select(ReceivedCheck, BusinessPartner, collection_account_alias, collection_bank_alias)
        .join(BusinessPartner, ReceivedCheck.customer_id == BusinessPartner.id)
        .outerjoin(collection_account_alias, ReceivedCheck.collection_bank_account_id == collection_account_alias.id)
        .outerjoin(collection_bank_alias, collection_account_alias.bank_id == collection_bank_alias.id)
        .order_by(ReceivedCheck.due_date, ReceivedCheck.id)
    )

    for check, customer, collection_account, collection_bank in session.execute(statement).all():
        rows.append(
            [
                check.id,
                check.check_number,
                customer.name,
                check.drawer_bank_name,
                check.drawer_branch_name,
                collection_bank.name if collection_bank else "",
                collection_account.account_name if collection_account else "",
                check.received_date,
                check.due_date,
                check.amount,
                check.currency_code.value,
                check.status.value,
                check.collected_transaction_id,
                check.reference_no,
                check.description,
            ]
        )

    _write_table(
        ws,
        [
            "ID",
            "Çek No",
            "Müşteri",
            "Çeki Veren Banka",
            "Çeki Veren Şube",
            "Tahsil Bankası",
            "Tahsil Hesabı",
            "Alış Tarihi",
            "Vade Tarihi",
            "Tutar",
            "Para Birimi",
            "Durum",
            "Tahsilat Hareket ID",
            "Referans No",
            "Açıklama",
        ],
        rows,
        money_columns=[10],
        date_columns=[8, 9],
        status_column=12,
    )


def export_financial_reports_to_excel(session: Session, *, as_of_date: date | None = None) -> Path:
    report_date = as_of_date or date.today()

    settings.export_folder.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = settings.export_folder / f"ftm_finansal_rapor_{timestamp}.xlsx"

    wb = Workbook()

    default_sheet = wb.active
    wb.remove(default_sheet)

    _create_dashboard_sheet(wb, session, report_date)
    _create_bank_balances_sheet(wb, session)
    _create_risk_sheets(wb, session, report_date)
    _create_transfer_recommendations_sheet(wb, session, report_date)
    _create_issued_checks_sheet(wb, session)
    _create_received_checks_sheet(wb, session)

    wb.active = 0
    wb.save(output_path)

    return output_path