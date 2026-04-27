from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.worksheet import Worksheet


@dataclass(frozen=True)
class FtmExcelMeta:
    title: str
    report_period: str
    created_by: str
    created_at: datetime


@dataclass(frozen=True)
class FtmExcelSummaryCard:
    title: str
    value: Any
    hint: str = ""
    card_type: str = "normal"
    number_format: str | None = None


class FtmExcelReportBuilder:
    def __init__(
        self,
        *,
        output_path: str | Path,
        meta: FtmExcelMeta,
    ) -> None:
        self.output_path = Path(output_path)
        self.meta = meta
        self.workbook = Workbook()

        default_sheet = self.workbook.active
        if default_sheet is not None:
            self.workbook.remove(default_sheet)

        self.colors = {
            "page_bg": "F8FAFC",
            "title": "0F172A",
            "subtitle": "64748B",
            "header_bg": "0F172A",
            "header_fg": "FFFFFF",
            "section_bg": "E0F2FE",
            "section_fg": "0F172A",
            "normal_bg": "FFFFFF",
            "muted_bg": "F8FAFC",
            "success_bg": "DCFCE7",
            "success_fg": "14532D",
            "warning_bg": "FEF3C7",
            "warning_fg": "78350F",
            "risk_bg": "FEE2E2",
            "risk_fg": "7F1D1D",
            "info_bg": "DBEAFE",
            "info_fg": "1E3A8A",
            "border": "CBD5E1",
            "soft_border": "E2E8F0",
            "total_bg": "EEF2FF",
        }

        self.font_name = "Calibri"
        self.money_format = '#,##0.00 [$₺-tr-TR]'
        self.number_format = '#,##0.00'
        self.integer_format = '#,##0'
        self.percent_format = '0.00%'
        self.date_format = 'dd.mm.yyyy'

    def add_sheet(
        self,
        title: str,
        *,
        freeze_panes: str | None = None,
        show_gridlines: bool = False,
    ) -> Worksheet:
        worksheet = self.workbook.create_sheet(title=title)
        worksheet.sheet_view.showGridLines = show_gridlines

        if freeze_panes:
            worksheet.freeze_panes = freeze_panes

        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_margins.left = 0.25
        worksheet.page_margins.right = 0.25
        worksheet.page_margins.top = 0.45
        worksheet.page_margins.bottom = 0.45
        worksheet.page_margins.header = 0.2
        worksheet.page_margins.footer = 0.2

        return worksheet

    def write_report_header(
        self,
        worksheet: Worksheet,
        *,
        last_column: int,
        subtitle: str | None = None,
    ) -> None:
        last_column_letter = get_column_letter(last_column)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_column)

        title_cell = worksheet["A1"]
        title_cell.value = self.meta.title
        title_cell.font = Font(name=self.font_name, size=18, bold=True, color=self.colors["title"])
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        period_cell = worksheet["A2"]
        period_cell.value = f"Rapor Dönemi: {self.meta.report_period}"
        period_cell.font = Font(name=self.font_name, size=10, color=self.colors["subtitle"])
        period_cell.alignment = Alignment(horizontal="left", vertical="center")

        info_parts = [
            f"Oluşturan: {self.meta.created_by}",
            f"Oluşturma Tarihi: {self.meta.created_at.strftime('%d.%m.%Y %H:%M')}",
        ]

        if subtitle:
            info_parts.append(subtitle)

        info_cell = worksheet["A3"]
        info_cell.value = " | ".join(info_parts)
        info_cell.font = Font(name=self.font_name, size=10, color=self.colors["subtitle"])
        info_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row_number in (1, 2, 3):
            worksheet.row_dimensions[row_number].height = 24

        for cell in worksheet[f"A1:{last_column_letter}3"][0]:
            cell.fill = PatternFill("solid", fgColor=self.colors["page_bg"])

    def write_section_title(
        self,
        worksheet: Worksheet,
        *,
        row: int,
        title: str,
        start_column: int,
        end_column: int,
    ) -> None:
        worksheet.merge_cells(start_row=row, start_column=start_column, end_row=row, end_column=end_column)

        cell = worksheet.cell(row=row, column=start_column)
        cell.value = title
        cell.font = Font(name=self.font_name, size=12, bold=True, color=self.colors["section_fg"])
        cell.fill = PatternFill("solid", fgColor=self.colors["section_bg"])
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = self._thin_border()

        worksheet.row_dimensions[row].height = 22

    def write_summary_cards(
        self,
        worksheet: Worksheet,
        *,
        start_row: int,
        start_column: int,
        cards: list[FtmExcelSummaryCard],
        columns: int = 4,
        card_width: int = 3,
    ) -> int:
        if columns <= 0:
            columns = 4

        row = start_row
        column = start_column

        for index, card in enumerate(cards):
            card_start_column = column
            card_end_column = column + card_width - 1

            worksheet.merge_cells(
                start_row=row,
                start_column=card_start_column,
                end_row=row,
                end_column=card_end_column,
            )
            worksheet.merge_cells(
                start_row=row + 1,
                start_column=card_start_column,
                end_row=row + 1,
                end_column=card_end_column,
            )
            worksheet.merge_cells(
                start_row=row + 2,
                start_column=card_start_column,
                end_row=row + 2,
                end_column=card_end_column,
            )

            title_cell = worksheet.cell(row=row, column=card_start_column)
            value_cell = worksheet.cell(row=row + 1, column=card_start_column)
            hint_cell = worksheet.cell(row=row + 2, column=card_start_column)

            title_cell.value = card.title
            value_cell.value = card.value
            hint_cell.value = card.hint

            card_colors = self._card_colors(card.card_type)

            for current_row in range(row, row + 3):
                for current_column in range(card_start_column, card_end_column + 1):
                    cell = worksheet.cell(row=current_row, column=current_column)
                    cell.fill = PatternFill("solid", fgColor=card_colors["bg"])
                    cell.border = self._thin_border()
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            title_cell.font = Font(name=self.font_name, size=9, bold=True, color=card_colors["fg"])
            value_cell.font = Font(name=self.font_name, size=13, bold=True, color=self.colors["title"])
            hint_cell.font = Font(name=self.font_name, size=8, color=self.colors["subtitle"])

            if card.number_format:
                value_cell.number_format = card.number_format

            column += card_width + 1

            if (index + 1) % columns == 0:
                row += 4
                column = start_column

        end_row = row + 3
        worksheet.row_dimensions[start_row].height = 20
        worksheet.row_dimensions[start_row + 1].height = 26
        worksheet.row_dimensions[start_row + 2].height = 28

        return end_row

    def write_key_value_table(
        self,
        worksheet: Worksheet,
        *,
        start_row: int,
        start_column: int,
        title: str,
        rows: list[tuple[str, Any, str | None]],
    ) -> int:
        self.write_section_title(
            worksheet,
            row=start_row,
            title=title,
            start_column=start_column,
            end_column=start_column + 2,
        )

        current_row = start_row + 1

        for label, value, number_format in rows:
            label_cell = worksheet.cell(row=current_row, column=start_column)
            value_cell = worksheet.cell(row=current_row, column=start_column + 1)

            label_cell.value = label
            value_cell.value = value

            label_cell.font = Font(name=self.font_name, size=10, bold=True, color=self.colors["title"])
            value_cell.font = Font(name=self.font_name, size=10, color=self.colors["title"])

            if number_format:
                value_cell.number_format = number_format

            for column in range(start_column, start_column + 2):
                cell = worksheet.cell(row=current_row, column=column)
                cell.border = self._thin_border()
                cell.fill = PatternFill("solid", fgColor=self.colors["normal_bg"])
                cell.alignment = Alignment(horizontal="left", vertical="center")

            current_row += 1

        return current_row + 1

    def write_table(
        self,
        worksheet: Worksheet,
        *,
        start_row: int,
        start_column: int,
        title: str,
        headers: list[str],
        rows: list[list[Any]],
        number_formats: dict[int, str] | None = None,
        formula_columns: dict[int, str] | None = None,
        row_styles: list[str] | None = None,
        auto_filter: bool = True,
        freeze_panes: str | None = None,
    ) -> tuple[int, int]:
        number_formats = number_formats or {}
        formula_columns = formula_columns or {}
        row_styles = row_styles or []

        end_column = start_column + len(headers) - 1

        self.write_section_title(
            worksheet,
            row=start_row,
            title=title,
            start_column=start_column,
            end_column=end_column,
        )

        header_row = start_row + 1
        data_start_row = header_row + 1

        for index, header in enumerate(headers, start=start_column):
            cell = worksheet.cell(row=header_row, column=index)
            cell.value = header
            cell.font = Font(name=self.font_name, size=10, bold=True, color=self.colors["header_fg"])
            cell.fill = PatternFill("solid", fgColor=self.colors["header_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = self._thin_border()

        for row_offset, row_values in enumerate(rows):
            current_row = data_start_row + row_offset
            row_style = row_styles[row_offset] if row_offset < len(row_styles) else "NORMAL"
            fill_color = self._row_fill(row_style)

            for column_offset, value in enumerate(row_values):
                current_column = start_column + column_offset
                cell = worksheet.cell(row=current_row, column=current_column)

                if column_offset in formula_columns:
                    cell.value = formula_columns[column_offset].format(row=current_row)
                else:
                    cell.value = value

                cell.font = Font(name=self.font_name, size=10, color=self.colors["title"])
                cell.fill = PatternFill("solid", fgColor=fill_color)
                cell.border = self._thin_border()
                cell.alignment = self._default_alignment(value)

                if column_offset in number_formats:
                    cell.number_format = number_formats[column_offset]

        last_data_row = data_start_row + max(len(rows) - 1, 0)

        if rows:
            total_row = last_data_row + 1

            label_cell = worksheet.cell(row=total_row, column=start_column)
            label_cell.value = "TOPLAM"
            label_cell.font = Font(name=self.font_name, size=10, bold=True, color=self.colors["title"])

            for current_column in range(start_column, end_column + 1):
                cell = worksheet.cell(row=total_row, column=current_column)
                cell.fill = PatternFill("solid", fgColor=self.colors["total_bg"])
                cell.border = self._thin_border()
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.font = Font(name=self.font_name, size=10, bold=True, color=self.colors["title"])

            for column_offset, number_format in number_formats.items():
                column_index = start_column + column_offset
                column_letter = get_column_letter(column_index)

                if number_format in {self.money_format, self.number_format, self.integer_format}:
                    total_cell = worksheet.cell(row=total_row, column=column_index)
                    total_cell.value = f"=SUBTOTAL(109,{column_letter}{data_start_row}:{column_letter}{last_data_row})"
                    total_cell.number_format = number_format

            end_row = total_row
        else:
            empty_row = data_start_row
            worksheet.cell(row=empty_row, column=start_column).value = "Kayıt bulunamadı."
            worksheet.cell(row=empty_row, column=start_column).font = Font(
                name=self.font_name,
                size=10,
                italic=True,
                color=self.colors["subtitle"],
            )
            end_row = empty_row

        if auto_filter and rows:
            worksheet.auto_filter.ref = (
                f"{get_column_letter(start_column)}{header_row}:"
                f"{get_column_letter(end_column)}{last_data_row}"
            )

        if freeze_panes:
            worksheet.freeze_panes = freeze_panes

        for row_number in range(header_row, end_row + 1):
            worksheet.row_dimensions[row_number].height = 22

        return end_row + 2, end_column

    def apply_conditional_percent_scale(
        self,
        worksheet: Worksheet,
        *,
        range_address: str,
        warning_threshold: float,
        risk_threshold: float,
    ) -> None:
        worksheet.conditional_formatting.add(
            range_address,
            CellIsRule(
                operator="greaterThanOrEqual",
                formula=[str(risk_threshold)],
                fill=PatternFill("solid", fgColor=self.colors["risk_bg"]),
            ),
        )
        worksheet.conditional_formatting.add(
            range_address,
            CellIsRule(
                operator="between",
                formula=[str(warning_threshold), str(risk_threshold - 0.000001)],
                fill=PatternFill("solid", fgColor=self.colors["warning_bg"]),
            ),
        )

    def set_column_widths(
        self,
        worksheet: Worksheet,
        widths: dict[int, float],
    ) -> None:
        for column_index, width in widths.items():
            column_letter = get_column_letter(column_index)
            dimension: ColumnDimension = worksheet.column_dimensions[column_letter]
            dimension.width = width

    def normalize_sheet(
        self,
        worksheet: Worksheet,
        *,
        max_column: int,
        default_row_height: float = 20,
    ) -> None:
        for row_number in range(1, worksheet.max_row + 1):
            if worksheet.row_dimensions[row_number].height is None:
                worksheet.row_dimensions[row_number].height = default_row_height

        for column_index in range(1, max_column + 1):
            column_letter = get_column_letter(column_index)
            if worksheet.column_dimensions[column_letter].width is None:
                worksheet.column_dimensions[column_letter].width = 14

    def save(self) -> str:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_path)

        return str(self.output_path)

    def _thin_border(self) -> Border:
        side = Side(style="thin", color=self.colors["soft_border"])

        return Border(left=side, right=side, top=side, bottom=side)

    def _card_colors(self, card_type: str) -> dict[str, str]:
        normalized_type = str(card_type or "normal").strip().lower()

        if normalized_type == "success":
            return {"bg": self.colors["success_bg"], "fg": self.colors["success_fg"]}

        if normalized_type == "warning":
            return {"bg": self.colors["warning_bg"], "fg": self.colors["warning_fg"]}

        if normalized_type == "risk":
            return {"bg": self.colors["risk_bg"], "fg": self.colors["risk_fg"]}

        if normalized_type == "info":
            return {"bg": self.colors["info_bg"], "fg": self.colors["info_fg"]}

        return {"bg": self.colors["muted_bg"], "fg": self.colors["title"]}

    def _row_fill(self, row_style: str) -> str:
        normalized_style = str(row_style or "NORMAL").strip().upper()

        if normalized_style in {"SUCCESS", "RECEIVED"}:
            return self.colors["success_bg"]

        if normalized_style in {"WARNING", "PROBLEM"}:
            return self.colors["warning_bg"]

        if normalized_style in {"RISK", "ISSUED"}:
            return self.colors["risk_bg"]

        if normalized_style in {"MUTED", "CLOSED"}:
            return self.colors["muted_bg"]

        return self.colors["normal_bg"]

    def _default_alignment(self, value: Any) -> Alignment:
        if isinstance(value, int | float | Decimal):
            return Alignment(horizontal="right", vertical="center")

        if isinstance(value, date | datetime):
            return Alignment(horizontal="center", vertical="center")

        return Alignment(horizontal="left", vertical="center", wrap_text=True)